"""
批量回测脚本：交叉测试8个模型目录 × 2个参数文件 = 16个组合
只回测15m的所有币种，生成Excel报告，标注组合名称

用法：
    python scripts/backtest_batch_models.py
    python scripts/backtest_batch_models.py --initial-capital 400 --test-days 90
"""

import json
import sys
import os
import argparse
import warnings
import signal
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
import subprocess
from concurrent.futures import ProcessPoolExecutor, as_completed

import pandas as pd
import numpy as np

# 抑制 scikit-learn 版本不匹配警告（不影响主要功能）
warnings.filterwarnings('ignore', message='.*InconsistentVersionWarning.*')
warnings.filterwarnings('ignore', category=UserWarning, module='sklearn')

# 添加项目根目录到 path
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.backtest_simulation import run_full_backtest, generate_summary_table
from src.python.data_fetcher import SYMBOLS, TIMEFRAMES

# 8个模型目录（根据文档）
MODEL_DIRS = [
    "data/models",
    "data/models_A",
    "data/models_B",
    "data/models_C",
    "data/models_v4",
    "data/models_A_v4",
    "data/models_B_old",
    "data/models_C_old",
]

# 2个参数文件
PARAMS_FILES = [
    "lightgbm_params_optuna.json",
    "lightgbm_params_optuna_v4.json",
]

# 只回测15m
TIMEFRAME_FILTER = "15m"


def load_prob_thresholds(config_path: Path) -> Dict[str, float]:
    """加载置信度阈值配置"""
    if not config_path.exists():
        return {}
    
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    return config.get("model_thresholds", {})


def load_backtest_config(config_path: Path) -> Tuple[Dict[str, float], Dict[str, bool]]:
    """加载回测配置（阈值和启用模型配置）"""
    if not config_path.exists():
        return {}, {}
    
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    model_thresholds = config.get("model_thresholds", {})
    enabled_models = config.get("enabled", {})
    
    return model_thresholds, enabled_models


def get_model_name(models_dir: str) -> str:
    """从模型目录路径提取模型名称"""
    # 例如: data/models -> models, data/models_B -> models_B
    return Path(models_dir).name


def get_params_name(params_file: str) -> str:
    """从参数文件名提取名称"""
    # 例如: lightgbm_params_optuna.json -> optuna
    # lightgbm_params_optuna_v4.json -> optuna_v4
    name = Path(params_file).stem
    if name.startswith("lightgbm_params_"):
        return name.replace("lightgbm_params_", "")
    return name


def run_backtest_for_combination(
    models_dir: str,
    params_file: str,
    initial_capital: float,
    test_days: int,
    prob_threshold: float,
    models_thresholds: Dict[str, float],
    n_jobs: int = 1,
) -> Dict[str, Any]:
    """
    运行单个组合的回测
    
    注意：参数文件在训练时使用，回测时模型已经训练好，所以这里主要是记录组合信息
    实际的回测使用已训练好的模型，不重新训练
    
    注意：组合名称会自动包含置信度阈值（格式：model+params+0.53），以便区分不同阈值的结果
    """
    model_name = get_model_name(models_dir)
    params_name = get_params_name(params_file)
    
    # 获取该模型的置信度阈值（如果有配置）
    model_threshold = models_thresholds.get(model_name, prob_threshold)
    
    # 在组合名称中包含阈值信息，以便区分不同阈值的结果（去重时能正确区分）
    # 格式：models_A+optuna+0.53 或 models_A+optuna+0.51
    combination_name = f"{model_name}+{params_name}+{model_threshold:.2f}"
    
    print(f"\n{'='*80}")
    print(f"📊 回测组合: {combination_name}")
    print(f"   模型目录: {models_dir}")
    print(f"   参数文件: {params_file}")
    print(f"   置信度阈值: {model_threshold*100:.1f}%")
    print(f"{'='*80}")
    
    models_dir_path = PROJECT_ROOT / models_dir
    if not models_dir_path.exists():
        print(f"⚠️  警告: 模型目录不存在: {models_dir_path}")
        return {
            "combination": combination_name,
            "model_dir": models_dir,
            "params_file": params_file,
            "error": "模型目录不存在",
            "results": [],
        }
    
    # 运行回测（只回测15m）
    try:
        # 如果并行处理有权限问题，自动降级为串行
        try:
            backtest_data = run_full_backtest(
                initial_capital=initial_capital,
                bet_ratio=0.05,  # 默认5%
                test_days=test_days,
                prob_threshold=model_threshold,
                primary_only=False,
                fee_rate=0.001,
                slippage=0.001,
                max_trades_per_day=20,
                models_dir=models_dir_path,
                stop_loss_pct=None,  # 批量回测不使用止损，便于对比
                consecutive_loss_limit=None,
                risk_pause_k_bars=None,
                n_jobs=n_jobs,
                timeframes=[TIMEFRAME_FILTER],  # 只回测15m
            )
        except PermissionError:
            # 权限问题，降级为串行
            print(f"⚠️  并行处理权限不足，改为串行处理...")
            backtest_data = run_full_backtest(
                initial_capital=initial_capital,
                bet_ratio=0.05,
                test_days=test_days,
                prob_threshold=model_threshold,
                primary_only=False,
                fee_rate=0.001,
                slippage=0.001,
                max_trades_per_day=20,
                models_dir=models_dir_path,
                stop_loss_pct=None,
                consecutive_loss_limit=None,
                risk_pause_k_bars=None,
                n_jobs=1,  # 串行
                timeframes=[TIMEFRAME_FILTER],  # 只回测15m
            )
        
        # 过滤只保留15m的结果（双重保险）
        filtered_results = []
        for r in backtest_data.get("results", []):
            if r.get("timeframe") == TIMEFRAME_FILTER:
                filtered_results.append(r)
        
        # 计算汇总统计
        if filtered_results:
            total_trades = sum(r.get("total_trades", 0) for r in filtered_results)
            total_wins = sum(r.get("wins", 0) for r in filtered_results)
            overall_win_rate = (total_wins / total_trades * 100) if total_trades > 0 else 0
            
            # 计算加权平均盈亏（按交易次数加权）
            total_profit = 0
            total_weight = 0
            for r in filtered_results:
                trades = r.get("total_trades", 0)
                if trades > 0:
                    total_profit += r.get("profit_pct", 0) * trades
                    total_weight += trades
            avg_profit = (total_profit / total_weight) if total_weight > 0 else 0
            
            # 计算最终资金总和
            final_capital_sum = sum(r.get("final_capital", initial_capital) for r in filtered_results)
            
            # 计算最大回撤（取最大值）
            # 注意：backtest_simulation.py返回的max_drawdown已经是百分比（乘以100），所以直接取最大值即可
            # 但需要确保值在合理范围内（0-100），如果超过100可能是计算错误
            max_drawdown_values = [r.get("max_drawdown", 0) for r in filtered_results]
            max_drawdown = max(max_drawdown_values, default=0)
            # 如果最大回撤超过100%，可能是数据异常，限制在100%
            if max_drawdown > 100:
                print(f"⚠️  警告: 检测到异常的最大回撤值 {max_drawdown}%，已限制为100%")
                max_drawdown = 100.0
        else:
            total_trades = 0
            total_wins = 0
            overall_win_rate = 0
            avg_profit = 0
            final_capital_sum = initial_capital * len(filtered_results) if filtered_results else initial_capital
            max_drawdown = 0
        
        return {
            "combination": combination_name,
            "model_dir": models_dir,
            "params_file": params_file,
            "prob_threshold": model_threshold,
            "summary": {
                "total_trades": total_trades,
                "total_wins": total_wins,
                "overall_win_rate": round(overall_win_rate, 2),
                "avg_profit_pct": round(avg_profit, 2),
                "final_capital_sum": round(final_capital_sum, 2),
                "max_drawdown": round(max_drawdown, 2),
            },
            "results": filtered_results,
            "config": backtest_data.get("config", {}),
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "combination": combination_name,
            "model_dir": models_dir,
            "params_file": params_file,
            "error": str(e),
            "results": [],
        }


def generate_excel_report(
    all_results: List[Dict[str, Any]],
    output_path: Path,
    initial_capital: float,
    merge_existing: bool = True,
):
    """
    生成Excel报告，按排名排序
    
    参数:
        all_results: 本次回测结果
        output_path: Excel文件路径
        initial_capital: 初始资金
        merge_existing: 是否合并已存在的数据（默认True，合并到同一个表格）
    """
    
    # 如果启用合并，尝试读取已存在的Excel文件
    existing_ranking = pd.DataFrame()
    existing_summary = pd.DataFrame()
    
    if merge_existing and output_path.exists():
        try:
            print(f"📖 读取已存在的Excel文件: {output_path.name}")
            # 读取已存在的排名表
            if '排名表' in pd.ExcelFile(output_path).sheet_names:
                existing_ranking = pd.read_excel(output_path, sheet_name='排名表')
                print(f"   找到 {len(existing_ranking)} 条已有排名数据")
            # 读取已存在的汇总表
            if '组合汇总' in pd.ExcelFile(output_path).sheet_names:
                existing_summary = pd.read_excel(output_path, sheet_name='组合汇总')
                print(f"   找到 {len(existing_summary)} 条已有汇总数据")
        except Exception as e:
            print(f"⚠️  读取已存在文件时出错: {e}，将创建新文件")
            existing_ranking = pd.DataFrame()
            existing_summary = pd.DataFrame()
    
    # 准备汇总数据
    summary_rows = []
    
    for result in all_results:
        if "error" in result:
            summary_rows.append({
                "排名": None,
                "组合名称": result["combination"],
                "模型目录": result["model_dir"],
                "参数文件": result["params_file"],
                "置信度阈值": result.get("prob_threshold", 0) * 100,
                "总交易次数": 0,
                "总胜场": 0,
                "总败场": 0,
                "总体胜率%": 0,
                "平均盈亏%": 0,
                "最终资金总和": initial_capital,
                "最大回撤%": 0,
                "状态": f"错误: {result['error']}",
            })
            continue
        
        summary = result.get("summary", {})
        results = result.get("results", [])
        
        total_trades = summary.get("total_trades", 0)
        total_wins = summary.get("total_wins", 0)
        total_losses = total_trades - total_wins
        
        summary_rows.append({
            "排名": None,  # 稍后排序后填充
            "组合名称": result["combination"],
            "模型目录": result["model_dir"],
            "参数文件": result["params_file"],
            "置信度阈值": result.get("prob_threshold", 0) * 100,
            "总交易次数": total_trades,
            "总胜场": total_wins,
            "总败场": total_losses,
            "总体胜率%": summary.get("overall_win_rate", 0),
            "平均盈亏%": summary.get("avg_profit_pct", 0),
            "最终资金总和": summary.get("final_capital_sum", initial_capital),
            "最大回撤%": summary.get("max_drawdown", 0),
            "状态": "成功",
        })
    
    # 创建DataFrame
    df_summary_new = pd.DataFrame(summary_rows)
    
    # 合并新旧汇总数据（如果启用合并）
    if merge_existing and not existing_summary.empty and not df_summary_new.empty:
        # 确保列顺序一致
        if set(df_summary_new.columns) == set(existing_summary.columns):
            # 合并数据（去重：如果组合名称相同，保留新的）
            df_summary_combined = pd.concat([existing_summary, df_summary_new], ignore_index=True)
            # 去重：保留最新的（按组合名称）
            df_summary_combined = df_summary_combined.drop_duplicates(
                subset=["组合名称"], 
                keep="last"
            )
            print(f"   合并后共 {len(df_summary_combined)} 条汇总数据（原有 {len(existing_summary)} + 新增 {len(df_summary_new)}）")
        else:
            # 列不一致，只使用新数据
            print(f"⚠️  汇总表列结构不一致，只使用新数据")
            df_summary_combined = df_summary_new
    elif merge_existing and not existing_summary.empty:
        # 只有旧数据，没有新数据
        df_summary_combined = existing_summary
    else:
        # 只有新数据
        df_summary_combined = df_summary_new
    
    # 对所有汇总数据重新排名（按总体胜率降序）
    df_summary = df_summary_combined.copy()
    
    # 过滤掉有错误的行进行排序
    df_valid = df_summary[df_summary["状态"] == "成功"].copy()
    if not df_valid.empty:
        df_valid = df_valid.sort_values("总体胜率%", ascending=False).reset_index(drop=True)
        df_valid["排名"] = df_valid.index + 1
        
        # 合并回原DataFrame
        df_summary.loc[df_summary["状态"] == "成功", "排名"] = df_valid["排名"].values
    
    # 重新排序（有排名的在前，按排名排序；无排名的在后）
    df_summary = pd.concat([
        df_summary[df_summary["排名"].notna()].sort_values("排名"),
        df_summary[df_summary["排名"].isna()],
    ]).reset_index(drop=True)
    
    # 创建统一排名表（所有组合的所有币种，按胜率%排序）
    ranking_rows = []
    for result in all_results:
        if "error" in result:
            continue
        
        combination = result["combination"]
        results = result.get("results", [])
        
        for r in results:
            # max_drawdown和min_over_initial_pct已经是百分比（从backtest_simulation.py返回）
            # 理论上应该在合理范围内，但为了防御性编程，再次限制范围
            max_dd = r.get("max_drawdown", 0)
            if max_dd < 0 or max_dd > 100:
                # 如果值异常，说明计算有问题，限制在合理范围
                max_dd = max(0, min(100, max_dd))
            
            min_over_init = r.get("min_over_initial_pct", 100)
            if min_over_init < 0 or min_over_init > 1000:
                # 如果值异常，说明计算有问题，限制在合理范围
                min_over_init = max(0, min(1000, min_over_init))
            
            ranking_rows.append({
                "排名": None,  # 稍后排序后填充
                "币种": r.get("symbol", "").replace("/USDT", ""),
                "周期": r.get("timeframe", ""),
                "交易次数": r.get("total_trades", 0),
                "胜场": r.get("wins", 0),
                "败场": r.get("losses", 0),
                "胜率%": round(r.get("win_rate", 0) * 100, 2),
                "最终资金": round(r.get("final_capital", initial_capital), 2),
                "盈亏%": round(r.get("profit_pct", 0), 2),
                "最大回撤%": round(max_dd, 2),
                "最低/初始%": round(min_over_init, 2),
                "组合": combination,
            })
    
    # 创建DataFrame
    df_ranking_new = pd.DataFrame(ranking_rows)
    
    # 合并新旧数据（如果启用合并）
    if merge_existing and not existing_ranking.empty and not df_ranking_new.empty:
        # 确保列顺序一致
        if set(df_ranking_new.columns) == set(existing_ranking.columns):
            # 合并数据（去重：如果组合+币种+周期相同，保留新的）
            df_ranking_combined = pd.concat([existing_ranking, df_ranking_new], ignore_index=True)
            # 去重：保留最新的（按组合+币种+周期）
            df_ranking_combined = df_ranking_combined.drop_duplicates(
                subset=["组合", "币种", "周期"], 
                keep="last"
            )
            print(f"   合并后共 {len(df_ranking_combined)} 条数据（原有 {len(existing_ranking)} + 新增 {len(df_ranking_new)}）")
        else:
            # 列不一致，只使用新数据
            print(f"⚠️  列结构不一致，只使用新数据")
            df_ranking_combined = df_ranking_new
    elif merge_existing and not existing_ranking.empty:
        # 只有旧数据，没有新数据
        df_ranking_combined = existing_ranking
    else:
        # 只有新数据
        df_ranking_combined = df_ranking_new
    
    # 对所有数据重新排名（按胜率%排序）
    if not df_ranking_combined.empty:
        df_ranking_combined = df_ranking_combined.sort_values("胜率%", ascending=False).reset_index(drop=True)
        df_ranking_combined["排名"] = df_ranking_combined.index + 1
        # 重新排列列顺序
        df_ranking = df_ranking_combined[[
            "排名", "币种", "周期", "交易次数", "胜场", "败场", 
            "胜率%", "最终资金", "盈亏%", "最大回撤%", "最低/初始%", "组合"
        ]]
    else:
        df_ranking = pd.DataFrame()
    
    # 保留详细结果表（用于其他用途）
    df_detail = df_ranking.copy() if not df_ranking.empty else pd.DataFrame()
    
    # 尝试写入Excel，如果openpyxl不可用则使用CSV
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 统一排名表（所有组合的所有币种，按胜率%排序）
            if not df_ranking.empty:
                df_ranking.to_excel(writer, sheet_name='排名表', index=False)
            
            # 汇总表（按组合汇总）
            df_summary.to_excel(writer, sheet_name='组合汇总', index=False)
            
            # 详细结果表（保留，用于其他用途）
            if not df_detail.empty:
                df_detail.to_excel(writer, sheet_name='详细结果', index=False)
            
            # 获取工作表对象以调整列宽（MacBook兼容格式）
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 定义表头样式（MacBook兼容）
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center_align = Alignment(horizontal="center", vertical="center")
            
            def format_worksheet(ws, df, sheet_name):
                """格式化工作表（MacBook兼容）"""
                # 调整列宽（支持超过26列）
                for idx, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(idx)
                    max_length = max(
                        df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                        len(str(col))
                    )
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                
                # 格式化表头（第一行）
                for idx, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(idx)
                    cell = ws[f"{col_letter}1"]
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                # 冻结首行（方便滚动查看）
                ws.freeze_panes = "A2"
                
                # 设置数字格式
                for row_idx in range(2, len(df) + 2):
                    for col_idx, col in enumerate(df.columns, start=1):
                        col_letter = get_column_letter(col_idx)
                        cell = ws[f"{col_letter}{row_idx}"]
                        
                        # 数字列右对齐
                        if col in ["排名", "交易次数", "胜场", "败场", "最终资金"]:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        # 百分比列右对齐
                        elif "%" in col:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            # 如果是百分比，确保显示为百分比格式
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "0.00%"
                        # 文本列左对齐
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 格式化各个工作表
            if not df_ranking.empty:
                worksheet_ranking = writer.sheets['排名表']
                format_worksheet(worksheet_ranking, df_ranking, '排名表')
            
            worksheet_summary = writer.sheets['组合汇总']
            format_worksheet(worksheet_summary, df_summary, '组合汇总')
            
            if not df_detail.empty:
                worksheet_detail = writer.sheets['详细结果']
                format_worksheet(worksheet_detail, df_detail, '详细结果')
        
        print(f"\n✅ Excel报告已生成: {output_path}")
    except (ModuleNotFoundError, ImportError):
        # openpyxl不可用，使用CSV
        csv_summary_path = output_path.with_suffix('.summary.csv')
        csv_detail_path = output_path.with_suffix('.detail.csv')
        
        # CSV 格式
        if not df_ranking.empty:
            csv_ranking_path = output_path.with_suffix('.ranking.csv')
            df_ranking.to_csv(csv_ranking_path, index=False, encoding='utf-8-sig')
        
        csv_summary_path = output_path.with_suffix('.summary.csv')
        df_summary.to_csv(csv_summary_path, index=False, encoding='utf-8-sig')
        
        if not df_detail.empty:
            csv_detail_path = output_path.with_suffix('.detail.csv')
            df_detail.to_csv(csv_detail_path, index=False, encoding='utf-8-sig')
        
        print(f"\n⚠️  openpyxl未安装，已生成CSV报告:")
        if not df_ranking.empty:
            print(f"   排名表: {csv_ranking_path}")
        print(f"   组合汇总: {csv_summary_path}")
        if not df_detail.empty:
            print(f"   详细结果: {csv_detail_path}")
        print(f"\n   如需Excel格式，请运行: pip install openpyxl")


def main():
    parser = argparse.ArgumentParser(
        description="批量回测8个模型×2个参数文件=16个组合（只回测15m）"
    )
    parser.add_argument(
        "--initial-capital",
        type=float,
        default=400.0,
        help="初始资金（USDC），默认400"
    )
    parser.add_argument(
        "--test-days",
        type=int,
        default=90,
        help="回测天数，默认90"
    )
    parser.add_argument(
        "--prob-threshold",
        type=float,
        default=None,
        help="默认置信度阈值（如果配置文件中没有指定），默认从配置文件读取"
    )
    parser.add_argument(
        "--threshold-config",
        type=str,
        default="config/backtest_thresholds.json",
        help="置信度阈值配置文件路径，默认config/backtest_thresholds.json"
    )
    parser.add_argument(
        "--output-excel",
        type=str,
        default=None,
        help="Excel输出路径，默认data/reports/batch_backtest_YYYYMMDD_HHMMSS.xlsx"
    )
    parser.add_argument(
        "--n-jobs",
        type=int,
        default=1,
        help="并行进程数，默认1（串行），2=适度并行（推荐），-1=使用全部CPU核心（可能过热）"
    )
    parser.add_argument(
        "--threshold-range",
        action="store_true",
        help="批量测试阈值范围：除了models外的其他模型，从0.51到0.60，按0.01递增（共10个阈值）"
    )
    
    args = parser.parse_args()
    
    # 读取回测配置（阈值和启用模型配置）
    threshold_config_path = PROJECT_ROOT / args.threshold_config
    models_thresholds, enabled_models = load_backtest_config(threshold_config_path)
    
    # 读取默认阈值
    if args.prob_threshold is None:
        default_threshold_path = PROJECT_ROOT / "config" / "polymarket_threshold.json"
        if default_threshold_path.exists():
            with open(default_threshold_path, 'r', encoding='utf-8') as f:
                default_config = json.load(f)
                default_prob_threshold = default_config.get("prob_threshold", 0.65)
        else:
            default_prob_threshold = 0.65
    else:
        default_prob_threshold = args.prob_threshold
    
    # 过滤掉被禁用的模型和不存在的模型目录
    active_model_dirs = []
    disabled_model_names = []
    missing_model_dirs = []
    
    for models_dir in MODEL_DIRS:
        model_name = get_model_name(models_dir)
        models_dir_path = PROJECT_ROOT / models_dir
        
        # 检查模型目录是否存在
        if not models_dir_path.exists():
            missing_model_dirs.append(model_name)
            print(f"⚠️  模型目录不存在，跳过: {model_name} ({models_dir})")
            continue
        
        # 检查是否被禁用
        is_enabled = enabled_models.get(model_name, True)
        if is_enabled:
            active_model_dirs.append(models_dir)
        else:
            disabled_model_names.append(model_name)
            print(f"⏭️  跳过禁用模型: {model_name} ({models_dir})")
    
    print("="*80)
    print("📊 批量回测脚本启动")
    print("="*80)
    print(f"总模型目录数: {len(MODEL_DIRS)}")
    print(f"存在的模型目录数: {len(active_model_dirs) + len(disabled_model_names)}")
    print(f"启用模型数: {len(active_model_dirs)}")
    if missing_model_dirs:
        print(f"⚠️  不存在的模型目录: {missing_model_dirs}")
    if disabled_model_names:
        print(f"⏭️  禁用模型: {disabled_model_names}")
    print(f"参数文件数: {len(PARAMS_FILES)}")
    print(f"总组合数: {len(active_model_dirs) * len(PARAMS_FILES)}")
    print(f"只回测周期: {TIMEFRAME_FILTER}")
    print(f"初始资金: ${args.initial_capital}")
    print(f"回测天数: {args.test_days}")
    print(f"默认置信度阈值: {default_prob_threshold*100:.1f}%")
    
    # 确定要测试的阈值范围
    if args.threshold_range:
        # 批量测试：除了models外的其他模型，从0.51到0.60，按0.01递增
        threshold_values = [round(0.51 + i * 0.01, 2) for i in range(10)]  # 0.51 到 0.60
        print(f"\n🔢 批量阈值测试模式")
        print(f"   阈值范围: {threshold_values[0]} ~ {threshold_values[-1]} (共{len(threshold_values)}个阈值)")
        print(f"   适用模型: 除models外的所有模型（models保持0.92）")
        print(f"   每个阈值将测试: {len(active_model_dirs) * len(PARAMS_FILES)} 个组合")
        print(f"   总任务数: {len(threshold_values) * len(active_model_dirs) * len(PARAMS_FILES)} 个回测任务")
        print(f"\n   配置文件中的阈值（仅作为参考，实际测试时会覆盖）:")
        if models_thresholds:
            print(f"   {models_thresholds}")
    else:
        # 普通模式：使用配置文件中的阈值
        if models_thresholds:
            print(f"模型特定阈值: {models_thresholds}")
        threshold_values = [None]  # None表示使用配置文件中的值
    print("="*80)
    
    # 运行所有组合的回测（可能多次，如果启用阈值范围测试）
    all_results = []
    excel_path = None
    
    for threshold_val in threshold_values:
        # 如果启用阈值范围测试，更新models_thresholds
        if args.threshold_range and threshold_val is not None:
            # 为除了models外的所有模型设置当前阈值
            current_models_thresholds = models_thresholds.copy()
            for model_name in current_models_thresholds:
                if model_name != "models":  # 除了models，其他都使用当前阈值
                    current_models_thresholds[model_name] = threshold_val
            print(f"\n{'='*80}")
            print(f"📊 测试阈值: {threshold_val*100:.1f}% (适用于除models外的所有模型)")
            print(f"   当前使用的阈值配置:")
            for model_name, threshold in sorted(current_models_thresholds.items()):
                print(f"     {model_name}: {threshold*100:.1f}%")
            print(f"{'='*80}")
        else:
            current_models_thresholds = models_thresholds
        
        # 准备所有组合的任务
        all_tasks = []
        for models_dir in active_model_dirs:
            for params_file in PARAMS_FILES:
                model_name = get_model_name(models_dir)
                model_threshold = current_models_thresholds.get(model_name, default_prob_threshold)
                all_tasks.append((
                    models_dir,
                    params_file,
                    args.initial_capital,
                    args.test_days,
                    default_prob_threshold,
                    current_models_thresholds,
                    args.n_jobs,
                ))
        
        # 运行本次阈值下的所有组合
        batch_results = []
        
        # 并行运行所有组合（如果组合数 > 1）
        total_combinations = len(all_tasks)
        if total_combinations > 1:
            # 计算组合级别的并行数
            # 注意：每个组合内部也可能并行（n_jobs），所以要控制总进程数
            cpu_count = os.cpu_count() or 4
            
            # 如果组合内部并行，限制组合级别的并行数，避免总进程数过多
            if args.n_jobs > 1:
                # 每个组合内部会启动 n_jobs 个进程
                # 总进程数 = combination_parallel * n_jobs
                # 限制总进程数不超过 CPU 核心数的 80%（留余量，避免过载和温度过高）
                max_total_processes = int(cpu_count * 0.8)
                combination_parallel = min(total_combinations, max(1, max_total_processes // args.n_jobs))
                print(f"\n🚀 并行运行 {total_combinations} 个组合（使用 {combination_parallel} 个进程，每个组合内部 {args.n_jobs} 个进程）...")
                print(f"   总进程数约: {combination_parallel * args.n_jobs}（CPU核心数: {cpu_count}，限制在80%避免过载）")
            else:
                # 组合内部串行（n_jobs=1），可以更多组合并行
                # 但为了控制CPU使用率和温度，限制在CPU核心数的60%（留余量）
                max_combination_parallel = max(1, int(cpu_count * 0.6))
                combination_parallel = min(total_combinations, max_combination_parallel)
                print(f"\n🚀 并行运行 {total_combinations} 个组合（使用 {combination_parallel} 个进程，每个组合内部串行）...")
                print(f"   CPU核心数: {cpu_count}，组合并行数: {combination_parallel}（限制在60%避免过载）")
            
            # 全局变量用于信号处理
            executor_ref = None
            futures_ref = None
            shutdown_requested = False
        
            def signal_handler(signum, frame):
                """处理 Ctrl+C 信号，优雅关闭所有进程"""
                nonlocal shutdown_requested, executor_ref, futures_ref
                if shutdown_requested:
                    print("\n\n⚠️  强制终止中...")
                    if executor_ref:
                        executor_ref.shutdown(wait=False, cancel_futures=True)
                    sys.exit(1)
                
                shutdown_requested = True
                print("\n\n🛑 收到停止信号，正在优雅关闭所有进程...")
                print("   请稍候，正在终止所有子进程...")
                
                if executor_ref and futures_ref:
                    # 取消所有未完成的任务
                    for future in futures_ref:
                        if not future.done():
                            future.cancel()
                    
                    # 关闭执行器（等待当前任务完成或超时）
                    try:
                        executor_ref.shutdown(wait=True, cancel_futures=True)
                    except Exception as e:
                        print(f"⚠️  关闭执行器时出错: {e}")
                
                print("✅ 所有进程已终止")
                sys.exit(0)
            
            # 注册信号处理器
            signal.signal(signal.SIGINT, signal_handler)
            signal.signal(signal.SIGTERM, signal_handler)
            
            try:
                with ProcessPoolExecutor(max_workers=combination_parallel) as executor:
                    executor_ref = executor
                    futures = {
                        executor.submit(run_backtest_for_combination, *task): task 
                        for task in all_tasks
                    }
                    futures_ref = list(futures.keys())
                    
                    completed = 0
                    for future in as_completed(futures):
                        if shutdown_requested:
                            print("\n⚠️  停止信号已收到，跳过剩余任务")
                            break
                        
                        try:
                            result = future.result(timeout=None)
                            batch_results.append(result)
                            completed += 1
                            print(f"✅ 完成组合 {completed}/{total_combinations}: {result.get('combination', '未知')}")
                        except KeyboardInterrupt:
                            print("\n\n🛑 收到中断信号")
                            shutdown_requested = True
                            break
                        except Exception as e:
                            if shutdown_requested:
                                break
                            task = futures[future]
                            model_name = get_model_name(task[0])
                            params_name = get_params_name(task[1])
                            combination_name = f"{model_name}+{params_name}"
                            error_msg = str(e)
                            # 如果是语法错误，提供更详细的提示
                            if "SyntaxError" in error_msg or "invalid syntax" in error_msg:
                                print(f"❌ 组合失败 {combination_name}: {error_msg}")
                                print(f"   ⚠️  检测到语法错误，请检查代码文件是否有问题")
                                print(f"   💡 建议：运行 'python3 -m py_compile scripts/backtest_batch_models.py' 检查语法")
                            else:
                                print(f"❌ 组合失败 {combination_name}: {error_msg}")
                            batch_results.append({
                                "combination": combination_name,
                                "model_dir": task[0],
                                "params_file": task[1],
                                "error": error_msg,
                                "results": [],
                            })
            except KeyboardInterrupt:
                print("\n\n🛑 收到中断信号，正在终止...")
                shutdown_requested = True
            finally:
                # 清理
                executor_ref = None
                futures_ref = None
        else:
            # 串行运行（如果只有一个组合）
            shutdown_requested = False
            
            def signal_handler(signum, frame):
                """处理 Ctrl+C 信号"""
                nonlocal shutdown_requested
                if shutdown_requested:
                    print("\n\n⚠️  强制终止中...")
                    sys.exit(1)
                
                shutdown_requested = True
                print("\n\n🛑 收到停止信号，正在终止...")
                sys.exit(0)
            
            signal.signal(signal.SIGINT, signal_handler)
            signal.signal(signal.SIGTERM, signal_handler)
            
            try:
                for task in all_tasks:
                    if shutdown_requested:
                        break
                    result = run_backtest_for_combination(*task)
                    batch_results.append(result)
            except KeyboardInterrupt:
                print("\n\n🛑 收到中断信号，正在终止...")
                shutdown_requested = True
        
        # 将本次批次的结果添加到总结果中
        all_results.extend(batch_results)
        
        # 每次运行后立即合并到Excel并重新排名
        if args.output_excel:
            excel_path = Path(args.output_excel)
        else:
            # 使用固定文件名，所有回测结果都保存在同一个文件中
            excel_path = PROJECT_ROOT / "data" / "reports" / "batch_backtest_all.xlsx"
        
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 合并模式：读取已存在的数据，合并后重新排名
        print(f"\n💾 保存结果到: {excel_path.name}")
        generate_excel_report(all_results, excel_path, args.initial_capital, merge_existing=True)
        
        # 如果还有更多阈值要测试，继续
        if args.threshold_range and threshold_val != threshold_values[-1]:
            print(f"\n⏭️  继续测试下一个阈值...\n")
    
    # 最终统计（所有阈值测试完成）
    if args.threshold_range:
        print(f"\n{'='*80}")
        print(f"✅ 所有阈值测试完成（共 {len(threshold_values)} 个阈值）")
        print(f"{'='*80}")
    
    # 打印最终汇总（如果不是阈值范围测试，或者所有阈值测试完成）
    if not args.threshold_range or threshold_val == threshold_values[-1]:
        print("\n" + "="*80)
        print("📊 批量回测完成")
        print("="*80)
        
        # 统计成功和失败的组合
        success_count = sum(1 for r in all_results if "error" not in r)
        error_count = len(all_results) - success_count
        
        print(f"总成功: {success_count}/{len(all_results)}")
        print(f"总失败: {error_count}/{len(all_results)}")
        
        if success_count > 0:
            # 找出最佳组合（按总体胜率）
            valid_results = [r for r in all_results if "error" not in r]
            best = max(valid_results, key=lambda x: x.get("summary", {}).get("overall_win_rate", 0))
            print(f"\n🥇 最佳组合: {best['combination']}")
            print(f"   总体胜率: {best['summary']['overall_win_rate']:.2f}%")
            print(f"   平均盈亏: {best['summary']['avg_profit_pct']:+.2f}%")
            print(f"   总交易次数: {best['summary']['total_trades']}")
        
        print(f"\n📄 Excel报告路径:")
        print(f"   {excel_path}")
        print(f"   完整路径: {excel_path.resolve()}")


if __name__ == "__main__":
    main()
