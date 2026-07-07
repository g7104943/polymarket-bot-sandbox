#!/usr/bin/env python3
"""
独立脚本：将 polymarket 各组合的 report_summary.txt 汇总到桌面，
支持 xlsx 或 Apple Numbers (.numbers)，每个组合一个工作表，排版清晰。
不依赖主系统，不嵌入主系统；始终覆盖同一文件（不生成新文件）。

与总日志一致：主系统（TypeScript）每小时整点写入各 logs_*/reports/report_summary.txt；
本脚本在每小时第 1 分钟（如 13:01、14:01）再读取并覆盖桌面同一份 xlsx/.numbers，避免 txt 尚未写入造成误差。
无需主系统「引用」本脚本——只要在项目根目录挂机运行 --watch 即可。

运行目录：在项目根目录（polyfun/）下执行即可，例如:
  cd /path/to/polyfun
  python polymarket/export_report_summaries_to_xlsx.py --watch --numbers   # 挂机，每小时覆盖同一 .numbers

用法:
  python polymarket/export_report_summaries_to_xlsx.py           # 导出一份 xlsx
  python polymarket/export_report_summaries_to_xlsx.py --numbers # 导出一份 .numbers
  python polymarket/export_report_summaries_to_xlsx.py --watch   # 每小时覆盖同一 xlsx
  python polymarket/export_report_summaries_to_xlsx.py --watch --numbers  # 每小时覆盖同一 .numbers（挂机）

依赖: pip install openpyxl（xlsx）；pip install numbers-parser（.numbers）
"""

import argparse
import sys
import time
from pathlib import Path
from datetime import datetime, timedelta

# 脚本所在目录 = polymarket/
SCRIPT_DIR = Path(__file__).resolve().parent
LOGS_DIR = SCRIPT_DIR
OUTPUT_XLSX = "polymarket_组合总日志.xlsx"
OUTPUT_NUMBERS = "polymarket_组合总日志.numbers"

# 主系统总汇总报告每小时整点更新；本脚本在每小时第 1 分钟读取，避免 txt 尚未写入
WATCH_OFFSET_MINUTES = 1  # 13:01, 14:01, 15:01 ...
WATCH_INTERVAL_SECONDS = 60 * 60


def _sheet_name_from_log_dir(dir_name: str) -> str:
    """logs_gru_btc_55 -> btc_55, logs_eth_10_90 -> eth_10_90, logs_btc -> btc"""
    name = dir_name.replace("logs_", "", 1)
    if name.startswith("gru_"):
        name = name[4:]
    return name[:31] if len(name) > 31 else name


def _sanitize_sheet_name(name: str, for_excel: bool = True) -> str:
    """Excel 工作表名不能含 \\ / * ? : [ ]"""
    for c in r'\/*?:[]':
        name = name.replace(c, "_")
    return name[:31] if len(name) > 31 else name


def _find_all_report_summaries():
    """返回 [(log_dir_name, report_summary_txt_path), ...] 按工作表名排序"""
    if not LOGS_DIR.exists():
        return []
    pairs = []
    for d in LOGS_DIR.iterdir():
        if not d.is_dir() or not d.name.startswith("logs_"):
            continue
        report = d / "reports" / "report_summary.txt"
        if report.exists():
            pairs.append((d.name, report))
    pairs.sort(key=lambda x: _sheet_name_from_log_dir(x[0]))
    return pairs


def _read_report_lines(path: Path) -> list[str]:
    with open(path, "r", encoding="utf-8") as f:
        return [line.rstrip("\n") for line in f.readlines()]


def _export_xlsx(pairs: list, out_path: Path, export_time: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    # 竖排：加粗、字号加大、居中
    cell_font = Font(bold=True, size=12)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, (log_dir_name, report_path) in enumerate(pairs):
        sheet_name = _sanitize_sheet_name(_sheet_name_from_log_dir(log_dir_name), for_excel=True)
        ws = wb.create_sheet(title=sheet_name, index=idx)
        lines = _read_report_lines(report_path)

        ws.cell(row=1, column=1, value=f"导出时间: {export_time}")
        ws.cell(row=1, column=1).font = cell_font
        ws.cell(row=1, column=1).alignment = cell_alignment
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=1).alignment = cell_alignment
        for row_idx, line in enumerate(lines, start=3):
            c = ws.cell(row=row_idx, column=1, value=line)
            c.font = cell_font
            c.alignment = cell_alignment
        ws.column_dimensions["A"].width = 80

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(out_path)


def _export_numbers(pairs: list, out_path: Path, export_time: str) -> None:
    from numbers_parser import Document

    doc = Document()
    sheet_names = [_sheet_name_from_log_dir(d[0]) for d in pairs]
    for c in r'\/*?:[]':
        sheet_names = [n.replace(c, "_")[:31] for n in sheet_names]

    for idx, (log_dir_name, report_path) in enumerate(pairs):
        sheet_name = sheet_names[idx]
        if idx == 0:
            sheet = doc.sheets[0]
            sheet.name = sheet_name
            table = sheet.tables[0]
        else:
            doc.add_sheet(sheet_name, "Table 1")
            sheet = doc.sheets[sheet_name]
            table = sheet.tables["Table 1"]

        lines = _read_report_lines(report_path)
        table.write(0, 0, f"导出时间: {export_time}")
        table.write(1, 0, "")
        for row_idx, line in enumerate(lines):
            table.write(2 + row_idx, 0, line)
        try:
            table.col_width(0, 400)
        except Exception:
            pass

    doc.save(str(out_path))


def do_export(use_numbers: bool) -> tuple[int, Path | None]:
    """执行一次导出。返回 (exit_code, out_path)。"""
    pairs = _find_all_report_summaries()
    if not pairs:
        print(f"未找到任何 report_summary.txt（在 {LOGS_DIR} 下 logs_*/reports/）")
        return 1, None

    desktop = Path.home() / "Desktop"
    desktop.mkdir(parents=True, exist_ok=True)
    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if use_numbers:
        try:
            out_path = desktop / OUTPUT_NUMBERS
            _export_numbers(pairs, out_path, export_time)
        except ImportError:
            print("导出 .numbers 需要: pip install numbers-parser")
            return 1, None
    else:
        try:
            out_path = desktop / OUTPUT_XLSX
            _export_xlsx(pairs, out_path, export_time)
        except ImportError:
            print("请先安装 openpyxl: pip install openpyxl")
            return 1, None

    print(f"已导出到: {out_path}（覆盖同一文件，不新建）")
    print(f"工作表: {[_sheet_name_from_log_dir(d[0]) for d in pairs]}")
    return 0, out_path


def main():
    parser = argparse.ArgumentParser(
        description="将各组合 report_summary 汇总到桌面 xlsx 或 Apple Numbers"
    )
    parser.add_argument(
        "--numbers",
        action="store_true",
        help="导出为 Apple Numbers (.numbers)，否则为 xlsx",
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help=f"按主系统总日志频率自动更新（每 {WATCH_INTERVAL_SECONDS // 60} 分钟一次）",
    )
    args = parser.parse_args()

    if args.watch:
        fmt = "Numbers" if args.numbers else "xlsx"
        print(f"自动更新已开启：每小时第 {WATCH_OFFSET_MINUTES} 分钟（如 13:01、14:01）读取 txt 并覆盖同一 {fmt} 文件")
        print("数据源: polymarket/logs_*/reports/report_summary.txt（晚 1 分钟读，避免主系统尚未写入）")
        print("按 Ctrl+C 停止\n")

        def next_run_time(now: datetime) -> datetime:
            """下次执行时刻：当前小时的 XX:01，若已过则下一小时 XX:01"""
            run_at = now.replace(minute=WATCH_OFFSET_MINUTES, second=0, microsecond=0)
            if run_at <= now:
                run_at += timedelta(hours=1)
            return run_at

        next_run = next_run_time(datetime.now())
        while True:
            now = datetime.now()
            if now < next_run:
                sleep_sec = (next_run - now).total_seconds()
                print(f"下次更新: {next_run.strftime('%H:%M')}（{int(sleep_sec)} 秒后）\n")
                try:
                    time.sleep(sleep_sec)
                except KeyboardInterrupt:
                    print("\n已停止自动更新")
                    return 0
            code, _ = do_export(use_numbers=args.numbers)
            if code != 0:
                return code
            next_run = next_run + timedelta(hours=1)
    else:
        code, _ = do_export(use_numbers=args.numbers)
        return code


if __name__ == "__main__":
    sys.exit(main())
